from variables import *
from classes.Resumen import Resumen
from typing import List
from variables import *
import os

# from openpyxl import load_workbook
# from jinja2 import Environment, FileSystemLoader
# from typing import List

from jinja2 import Environment, FileSystemLoader
import openpyxl
from openpyxl import load_workbook


class File:
    def get_dirs_of_dir_mes(dir_path, mes):
        return [file for file in os.listdir(dir_path) if os.path.isdir(os.path.join(dir_path, file)) and file[:3].lower() == mes.lower()]

    def get_files_of_dir(dir, filename):
        return [f for f in os.listdir(dir) if os.path.isfile(os.path.join(dir, f)) and f == filename]


    def is_jinja2_template(file_path):
        # Cargar el libro de trabajo XLSX
        workbook = load_workbook(file_path)

        # Recorrer todas las hojas del libro de trabajo
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Recorrer todas las celdas de la hoja
            for row in sheet.iter_rows():
                for cell in row:
                    # Comprobar si la celda contiene una variable de plantilla de Jinja2
                    if cell.data_type == 's' and '{{' in cell.value and '}}' in cell.value:
                        return True
        return False

    def generar_reporte_mensual(org: str, resumenes: List[Resumen]):
        template_path = rutaArchivo + "Plantillas" + '\\' + f"Reporte_Mensual_{org}_plantilla.xlsx"
        output_path = rutaArchivo + "Reportes" + '\\' + f"Reporte_Mensual_{org}.xlsx"

        if not os.path.exists(os.path.join(template_path)):
            print('No existe plantilla', template_path)
            return
        
        print('es plantilla', File.is_jinja2_template(template_path))

        # print(template_path)
        # print(output_path)
        i = 0
        context = []
        for item in resumenes:
            i+=1
            if i == 5:
                break
            context.append({
                "afiliado": item.afiliado,
                "terminal": item.terminal,
                "terminal": item.terminal,
                "comercio": item.comercio,
                "ubicacion": item.ubicacion,
                "doeContador": item.doeContador,
                "doeMonto": item.doeMonto,
                "dbncContador": item.doeMonto,
                "dbncMonto": item.dbncMonto,
                "dpContador": item.dpContador,
                "dpMonto": item.dpMonto,
                "voeContador": item.voeContador,
                "voeMonto": item.voeMonto,
                "vbncContador": item.vbncContador,
                "vbncMonto": item.vbncMonto,
                "moeContador": item.moeContador,
                "moeMonto": item.moeMonto,
                "mbncContador": item.mbncContador,
                "mbncMonto": item.mbncMonto
            })

        print('Len arr',len(context), org)

        archivos_en_ruta = os.listdir(rutaArchivo + "Plantillas" + '\\')
        archivo_buscado = f"Reporte_Mensual_{org}_plantilla.xlsx"
        if archivo_buscado in archivos_en_ruta:
            print('Existe el archivo')
        else:
            print('No Existe el archivo')

        # # Crear un entorno Jinja2
        # env = Environment(loader=FileSystemLoader('/'))

        # # Cargar la plantilla desde el archivo
        # template = env.get_template(template_path)

        # # crear un entorno Jinja2 y cargar la plantilla en la memoria
        # loader = FileSystemLoader(searchpath='/')
        # env = Environment(loader=loader)
        # template = env.from_string(worksheet)

        # renderizar la plantilla y escribir los resultados en el archivo de salida
        # with open(output_path, "wb") as output_file:
        #     rendered = template.render(context=context)
        #     output_file.write(rendered)



        # print('Crete reporte excel:', org, 'registros', len(resumenes))
        # inFile = f"Reporte_Mensual_{org}_plantilla.xlsx"
        # outFile = f"Reporte_Mensual_{org}.xlsx"
        # print(inFile)
        # with open(os.path.join(rutaArchivo, "Plantillas", inFile ), "rb"), \
        #     open(os.path.join(rutaArchivo, "Reportes", outFile), "wb") as out_file:
        #     print('bu0')
        #     context = {"resumenes": resumenes}
        #     print('bu1')
        #     template_loader = FileSystemLoader(searchpath=rutaArchivo)
        #     template_env = Environment(loader=template_loader)
        #     print('bu2')
        #     template = template_env.get_template("Plantillas/" + inFile).format(org)
        #     print('bu3')
        #     rendered_template = template.render(context)
        #     print('bu4')
        #     out_file.write(rendered_template)
